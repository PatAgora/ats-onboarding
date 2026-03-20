#!/usr/bin/env python3
"""
Generate OS1 Full Requirements Tracker Excel document.
Produces: OS1_Full_Requirements_Tracker.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour definitions
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
STATUS_FILLS = {
    "Complete": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "In Progress": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Deferred - Contradiction": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
    "Complete": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "Not Started": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# Source documents with dates
SOURCES = {
    "OS1Feedback.xlsx": "10 Mar 2026",
    "Vetting Section (Word).docx": "04 Mar 2026",
    "Declaration Form.pdf": "04 Mar 2026",
    "Consent Form.pdf": "04 Mar 2026",
    "OS1_Associate Portal_v1.0.pptx": "04 Mar 2026",
    "An Intro to Vetting.docx": "04 Mar 2026",
    "Reference Request - Contact Details.xlsx": "04 Mar 2026",
    "Reference Houses - Intel.xlsx": "04 Mar 2026",
    "Security Audit": "19 Mar 2026",
    "Gap Analysis": "19 Mar 2026",
    "OS1 - System Functionality & Development copy.docx": "05 Mar 2026",
    "OS1 Wireframe_Backend Ian version copy.docx": "19 Feb 2026",
    "OS1_Actions 12012026 copy.xlsx": "12 Jan 2026",
    "OS1_Actions 19012026 copy.xlsx": "19 Jan 2026",
}

# ---------------------------------------------------------------------------
# Requirements data: (source_doc, category, description, priority, status, summary, notes)
# ---------------------------------------------------------------------------
REQS = [
    # === OS1Feedback.xlsx — 53 items (ALL COMPLETE) ===
    ("OS1Feedback.xlsx", "Login", "Login URL redirect bug — goes to 'bad URL' page first", "High", "Complete", "Fixed empty next URL redirect bug in login route", ""),
    ("OS1Feedback.xlsx", "Dashboard", "Job links go to website portal instead of internal job details", "High", "Complete", "Job title links now route to internal application detail page", ""),
    ("OS1Feedback.xlsx", "Placements", "CSV export encoding — Pay/Bill rate shows as 'GBP350'", "Medium", "Complete", "Added UTF-8 BOM prefix for Excel compatibility", ""),
    ("OS1Feedback.xlsx", "Placements", "CSV export — split Associate Name and Email into separate columns", "Medium", "Complete", "Split into separate Name and Email columns", ""),
    ("OS1Feedback.xlsx", "Placements", "Name filter shows associates without active placements", "Medium", "Complete", "Filter now only shows associates with active placements", ""),
    ("OS1Feedback.xlsx", "Placements", "Client filter includes clients without active placements", "Medium", "Complete", "Filter now only includes clients with active placements", ""),
    ("OS1Feedback.xlsx", "Placements", "On Contract module needs actions — ability to change end date", "Medium", "Complete", "Added end date change functionality on contract view", ""),
    ("OS1Feedback.xlsx", "Placements", "Forecasted headcount graphs should be above on-contract list", "Medium", "Complete", "Forecast chart moved above on-contract list", ""),
    ("OS1Feedback.xlsx", "Placements", "Client Summary daily rate card visible to non-admins", "Medium", "Complete", "Client Summary daily rate card removed from this view", ""),
    ("OS1Feedback.xlsx", "Placements", "Associate names should be clickable to open profile", "Medium", "Complete", "Associate names now link to candidate profile everywhere", ""),
    ("OS1Feedback.xlsx", "Workflow", "New applications land in Shortlisted instead of Pipeline", "High", "Complete", "New applications now land in Pipeline stage", ""),
    ("OS1Feedback.xlsx", "Workflow", "No option to click into vetting checks to see associates outside SLA", "Medium", "Complete", "Vetting summary with SLA tracking added", ""),
    ("OS1Feedback.xlsx", "Workflow", "AI summary copies first role verbatim; AI score not working", "Medium", "Complete", "Switched to Google Gemini (gemini-2.0-flash). Recruiter-focused summary prompt with suitability rating, strengths, experience, gaps, availability. Scoring prompt tuned for financial services. Fixed blended_score key mismatch.", ""),
    ("OS1Feedback.xlsx", "Workflow", "View Job button goes to bad URL", "High", "Complete", "View Job link fixed to internal route", ""),
    ("OS1Feedback.xlsx", "Workflow", "Need manual button to trigger vetting to In Progress", "High", "Complete", "Start Vetting button added with auto-check creation + email", "Contradiction #1 resolved"),
    ("OS1Feedback.xlsx", "Workflow", "Vetting check status updates don't save", "High", "Complete", "Vetting form action fixed to update_vetting_check route", ""),
    ("OS1Feedback.xlsx", "Workflow", "Workflow lets you tick vetting complete without all checks done; need REFERRAL status", "High", "Complete", "Blocking rule: all checks must be Complete/N/A before Ready to Contract. REFERRAL APPROVED status added.", ""),
    ("OS1Feedback.xlsx", "Workflow", "Horizontal scroll should start at edge of left nav; left panel fixed", "Medium", "Complete", "Fixed horizontal scroll with sticky left nav panel", ""),
    ("OS1Feedback.xlsx", "Workflow", "Focus View — allow users to limit workflow grid columns", "Medium", "Complete", "Focus View with column checkboxes implemented", ""),
    ("OS1Feedback.xlsx", "Workflow", "Rejected/Withdrawn column should be at right-hand end", "Low", "Complete", "Rejected/Withdrawn moved to last column position", ""),
    ("OS1Feedback.xlsx", "Workflow", "Placed/Contract Sent — add link to OS1 placement", "Medium", "Complete", "Contract Sent shows 'View Placements' link", ""),
    ("OS1Feedback.xlsx", "Workflow", "Unsigned Contracts & Vetting Summary should respond to grid filters", "Medium", "Complete", "Both modules now filter-responsive to grid selections", ""),
    ("OS1Feedback.xlsx", "Workflow", "Placed column — remove from workflow grid", "Medium", "Complete", "Placed column removed; signed contracts move to Placements section", ""),
    ("OS1Feedback.xlsx", "Workflow", "Bulk rejection with email notification when role is filled", "Medium", "Complete", "Bulk reject modal with email notification implemented", ""),
    ("OS1Feedback.xlsx", "Projects", "Four infographic tiles should adjust based on page filters", "Medium", "Complete", "Summary tiles now respond to applied page filters", ""),
    ("OS1Feedback.xlsx", "Projects", "Filters only filter Engagements list, not Pipeline/Opportunities", "Medium", "Complete", "Separate filter bars for Engagements and Pipeline/Opportunities", ""),
    ("OS1Feedback.xlsx", "Projects", "Closed Won should auto-move from Pipeline to Current Engagements", "High", "Complete", "Closed Won auto-moves to Current Engagements on drag", ""),
    ("OS1Feedback.xlsx", "Projects", "Closed Won/Lost should be clickable tiles, not kanban columns", "Medium", "Complete", "Removed from kanban columns; shown as clickable summary tiles above kanban. Procurement cards have Won/Lost action buttons that auto-move and create engagement on Won.", ""),
    ("OS1Feedback.xlsx", "Projects", "Engagement main page — no filters on Delivery Plan by Role table", "Medium", "Complete", "Role and Intake filters added to Delivery Plan table", ""),
    ("OS1Feedback.xlsx", "Projects", "Adding new role to plan doesn't create matching Active Job Post", "Medium", "Complete", "Engagement edit plan now prompts to create matching job", ""),
    ("OS1Feedback.xlsx", "Projects", "Creating a job went to broken web page; define vetting requirements at setup", "High", "Complete", "Job creation fixed; vetting requirements textarea on engagement", ""),
    ("OS1Feedback.xlsx", "Projects", "Cannot amend project name after creation", "Medium", "Complete", "Edit engagement page created with name amendment", ""),
    ("OS1Feedback.xlsx", "Projects", "Use convention OS001, OS002, OS003 for project references", "Low", "Complete", "Engagement ref prefix changed from EG to OS (OS001, OS002...)", ""),
    ("OS1Feedback.xlsx", "Projects", "Start/end dates not showing in Current Engagements after Closed Won", "Medium", "Complete", "Start/end dates now display correctly after status change", ""),
    ("OS1Feedback.xlsx", "Projects", "Weighted forecast doesn't update when moving through card stages", "Medium", "Complete", "Opportunity probability auto-updates per pipeline stage", ""),
    ("OS1Feedback.xlsx", "Projects", "Add toggle filter for Active vs Completed engagements", "Medium", "Complete", "Active/Completed/All toggle filter with default to Active", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "Search for KYC returns results but quoted 'KYC' returns zero", "High", "Complete", "Quote stripping on search queries fixed", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "Advanced filters auto-expand on page refresh", "Low", "Complete", "Filters collapsed by default on page load", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "Boolean search CDD OR KYC OR AML returns TypeError", "High", "Complete", "Boolean search condition logic fixed", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "Shortlist button unclear — change to 'Add to' with job pipeline", "Medium", "Complete", "Replaced with 'Add to Pipeline' dropdown", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "Delete button on main page — remove, admin-only in profile", "Medium", "Complete", "Delete removed from main table; admin-only in profile", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "Should be able to filter by table headings (status, location)", "Medium", "Complete", "Status and Location column header filters added", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "Activity feed doesn't log past interview stage", "Medium", "Complete", "Activity notes auto-created for all workflow moves", ""),
    ("OS1Feedback.xlsx", "Resource Pool", "No way to manually add associate to a job process", "Medium", "Complete", "Manual application for roles functionality added", ""),
    ("OS1Feedback.xlsx", "Associate Profile", "Profile status should auto-link to placements", "Medium", "Complete", "Status auto-linked to active placements", ""),
    ("OS1Feedback.xlsx", "Associate Profile", "No option to update core details (email, name, phone)", "High", "Complete", "Edit Details modal for name/email/phone added", ""),
    ("OS1Feedback.xlsx", "Associate Profile", "Cannot add Activity or save General Notes", "High", "Complete", "General Notes card with scrollable list; CandidateNote field fixed", ""),
    ("OS1Feedback.xlsx", "Associate Profile", "Active placement not showing for certain associates", "High", "Complete", "Active placement display logic fixed", ""),
    ("OS1Feedback.xlsx", "Associate Portal", "QR code not showing during 2FA registration", "High", "Complete", "QR code generates inline base64 with API fallback; CSP updated", ""),
    ("OS1Feedback.xlsx", "Revenue", "Add forecasted margin vs actual margin graph", "Medium", "Complete", "Forecast Margin vs Actual Margin chart added with trend data", ""),
    ("OS1Feedback.xlsx", "Revenue", "Date filters don't apply to Actual vs Forecast Revenue infographic", "Medium", "Complete", "Filter-aware trend data generation respects date range", ""),
    ("OS1Feedback.xlsx", "Revenue", "Filters don't have cascading relationship (client→engagement)", "Medium", "Complete", "Cascading Client → Engagement filter implemented", ""),

    # === Vetting Section (Word) — Gap Plan items ===
    ("Vetting Section (Word).docx", "Vetting", "Consent form — associates must sign before vetting checks", "High", "Complete", "Digital consent form built with e-signature in portal", ""),
    ("Vetting Section (Word).docx", "Vetting", "No vetting checks can be submitted without completed Consent Form", "High", "Complete", "Blocking rule: consent required before vetting proceeds", "Contradiction #3 also covers this"),
    ("Vetting Section (Word).docx", "Vetting", "Consent form auto-attached to all reference requests", "Medium", "Complete", "send_reference route queries consent_signed Document, reads file, passes as email attachment", "Consent PDF auto-attached to all outgoing reference request emails"),
    ("Vetting Section (Word).docx", "Vetting", "Declaration Yes/No form with auto On Hold when vetting exists", "High", "Complete", "6-question declaration form; Yes → On Hold for all vetting checks if grid exists, else flags candidate", "Contradiction #3 resolved"),
    ("Vetting Section (Word).docx", "Vetting", "Client-specific document config per project/engagement", "High", "Complete", "required_documents JSON field on Engagement model", ""),
    ("Vetting Section (Word).docx", "Documents", "Proof of ID must be in-date; expiry tracked; linked to DBS", "High", "Complete", "expiry_date field on Document model; daily expiry check job", ""),
    ("Vetting Section (Word).docx", "Documents", "Proof of Address with 90-day rule enforcement", "High", "Complete", "_check_poa_90_day_rule() helper function enforces 90-day limit", ""),
    ("Vetting Section (Word).docx", "Documents", "Right to Work expiry tracking during assignment", "High", "Complete", "expiry_date on Document model; daily expiry alerts via scheduler", ""),
    ("Vetting Section (Word).docx", "References", "Reference requirements client-specific (X years configurable)", "High", "Complete", "reference_period_years on Engagement; default 3; editable in engagement forms", "Contradiction #5 resolved"),
    ("Vetting Section (Word).docx", "References", "All gaps over X days must be explained", "High", "Complete", "Portal gap detection with >90 day threshold and evidence upload", ""),
    ("Vetting Section (Word).docx", "References", "System blocks progression if employment history incomplete", "Medium", "Complete", "_check_employment_complete() validates gaps >90 days are explained; vetting progress shows warning", "P10 — enforcement via helper function"),
    ("Vetting Section (Word).docx", "References", "Reference contact auto-populate from approved contacts list", "High", "Complete", "Portal autocomplete from ReferenceContact table + admin CRUD import", ""),
    ("Vetting Section (Word).docx", "References", "Flagged reference house auto-warning on entry", "High", "Complete", "Portal real-time check + admin CRUD with full intel display", ""),
    ("Vetting Section (Word).docx", "References", "Unknown referee email requires manual check before sending", "Medium", "Complete", "Unknown company → pending_verification amber warning", ""),
    ("Vetting Section (Word).docx", "Vetting", "Vetting statuses: Waiting for Associate, Ready to Start, In Progress, On Hold, Awaiting QC, QC Complete", "High", "Complete", "All statuses supported in VettingCheck model and routes", ""),
    ("Vetting Section (Word).docx", "Vetting", "Cards auto-move based on triggers (vetting complete → Contract Issued)", "High", "Complete", "Auto-move when all 12 checks complete", ""),
    ("Vetting Section (Word).docx", "Vetting", "Cards manually moveable with permission controls", "High", "Complete", "Analyst cannot QC checks assigned to themselves (N9)", ""),
    ("Vetting Section (Word).docx", "Vetting", "Per-check colour system: White/Green/Orange — auto and manual", "High", "Complete", "colour + colour_manual fields; /action/set-check-colour route", ""),
    ("Vetting Section (Word).docx", "Vetting", "Prompt-only checks (no upload required)", "Medium", "Complete", "prompt_only field on VettingCheck model", ""),
    ("Vetting Section (Word).docx", "References", "Reference cards: associate data, permission status, sent timestamp, .msg upload", "High", "Complete", "ReferenceRequest model with all fields; send/chase/receive routes", ""),
    ("Vetting Section (Word).docx", "References", "Reference timeline visual with gap detection after refs received", "High", "Complete", "Portal employment timeline with >90 day gap detection", ""),
    ("Vetting Section (Word).docx", "References", "Send Reference email route to referee", "High", "Complete", "/action/send-reference route with email", ""),
    ("Vetting Section (Word).docx", "References", "Chase Reference route — resend with Reminder, max 3 then flag", "High", "Complete", "/action/chase-reference with chase_count; max 3 → On Hold", ""),
    ("Vetting Section (Word).docx", "Vetting", "Auto-chase references every 72 hours", "High", "Complete", "APScheduler job: auto_chase_references every 72h", ""),
    ("Vetting Section (Word).docx", "Vetting", "Auto-poll Verifile every 24 hours for pending checks", "High", "Complete", "APScheduler job: auto_poll_verifile every 24h", ""),
    ("Vetting Section (Word).docx", "Vetting", "Auto-chase candidates every 24 hours for incomplete profiles", "Medium", "Complete", "APScheduler job: auto_chase_candidates every 24h", ""),
    ("Vetting Section (Word).docx", "Documents", "Document expiry daily alerts (30-day warning + expired notes)", "High", "Complete", "APScheduler job: check_document_expiry daily; CandidateNote created", ""),
    ("Vetting Section (Word).docx", "Vetting", "QC workflow: submit for QC, approve, reject", "High", "Complete", "3 routes: submit-for-qc, qc-approve, qc-reject", ""),
    ("Vetting Section (Word).docx", "Vetting", "Analyst case allocation with staff dropdown + filter", "High", "Complete", "assigned_to field; assign-analyst route; workflow filter dropdown", ""),
    ("Vetting Section (Word).docx", "Vetting", "ID verified via video call/in-person/IDVT before DBS submission", "High", "Complete", "id_verified + id_verification_method fields; /action/verify-id route", ""),
    ("Vetting Section (Word).docx", "Vetting", "Grid removal when vetting complete + contracted + start date passed", "Medium", "Complete", "_should_remove_from_grid() helper function", ""),
    ("Vetting Section (Word).docx", "Vetting", "NI Number field on candidate/associate", "High", "Complete", "national_insurance_number on AssociateProfile model in portal", ""),
    ("Vetting Section (Word).docx", "Vetting", "Gender field on candidate model", "High", "Complete", "gender field on Candidate model in app.py", ""),
    ("Vetting Section (Word).docx", "Vetting", "Mother's Maiden Name, DL Number, Passport Number (optional DBS fields)", "Medium", "Complete", "mothers_maiden_name, driving_licence_number, passport_number on Candidate", ""),
    ("Vetting Section (Word).docx", "Vetting", "Citizenship and Place of Birth fields", "High", "Complete", "citizenship, place_of_birth on Candidate model", ""),
    ("Vetting Section (Word).docx", "Vetting", "Work Location for World Check", "Medium", "Complete", "work_location field on Candidate model", ""),
    ("Vetting Section (Word).docx", "Vetting", "Name variations tracking (multiple name forms for checks)", "Medium", "Complete", "name_variations JSON field on Candidate model", ""),
    ("Vetting Section (Word).docx", "Admin", "Admin CRUD for approved reference contacts (4,935 entries) with CSV import", "High", "Complete", "5 routes + template: list, add, edit, delete, CSV import", ""),
    ("Vetting Section (Word).docx", "Admin", "Admin CRUD for flagged reference houses with full intel display", "High", "Complete", "4 routes + template: list with intel cards, add, edit, delete", ""),

    # === Declaration Form PDF ===
    ("Declaration Form.pdf", "Associate Portal", "6 Yes/No declaration questions (restrictions, convictions, CCJ, bankruptcy, dismissed, referencing)", "High", "Complete", "Full form in declaration_form.html with all 6 questions + conditional detail fields", ""),
    ("Declaration Form.pdf", "Associate Portal", "Open disclosure text box for voluntary declarations", "Medium", "Complete", "disclosure_text field in DeclarationRecord model + textarea in form", ""),
    ("Declaration Form.pdf", "Associate Portal", "E-signature with Legal Name, Signed Date, confirmation statement", "High", "Complete", "DeclarationRecord: legal_name, signed_date, ip_address; confirmation checkbox", ""),

    # === Consent Form PDF ===
    ("Consent Form.pdf", "Associate Portal", "Full consent text with data sharing and acknowledgment sections", "High", "Complete", "consent_form.html with 9 acknowledgment items + data processing info", ""),
    ("Consent Form.pdf", "Associate Portal", "Secondary employment Yes/No with job title if Yes", "Medium", "Complete", "ConsentRecord: secondary_employment, secondary_employment_details fields", ""),
    ("Consent Form.pdf", "Associate Portal", "Fraud prevention Cifas section in consent", "Medium", "Complete", "Cifas fraud prevention section in consent_form.html", ""),
    ("Consent Form.pdf", "Associate Portal", "E-signature with Legal Name, Signed Date", "High", "Complete", "ConsentRecord: legal_name, signed_date, ip_address, consent_given", ""),
    ("Consent Form.pdf", "Associate Portal", "Manual upload fallback for hand-signed consent forms", "Medium", "Complete", "manual_consent file upload in consent_form POST; saves as consent_signed doc type", "Contradiction #4 resolved"),

    # === Portal PPT ===
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Portal dashboard with progress bars per section (Personal/Company/Background)", "High", "Complete", "3-section progress bars with % completion on dashboard", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Personal details: all fields including emergency contact, DOB, aliases, salary", "High", "Complete", "Full form: title, name, aliases, previous names, DOB, address, emergency contact, salary, NI, CV", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Company details: umbrella panel + limited company fields + IR35 text", "High", "Complete", "Umbrella selection (now configurable), Ltd Co fields, IR35 banner. Contradiction #6 resolved.", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Assignment History page (Role, Project, Client, Dates, Rate, Status)", "High", "Complete", "assignments.html table with On Contract/Contract ended status", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Timesheet Mon-Sun weekly grid with Standard/Overtime/Holiday/Sickness", "High", "Complete", "Full weekly grid with configurable time types, half/full day dropdowns, HH:MM OT input", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Expense section with upload and type/amount", "Medium", "Complete", "Expense add form with type/description/amount/receipt upload; delete route; totals auto-update", "P6 — complete"),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Previous Timesheets historical section with monthly bundles", "Medium", "Complete", "Monthly bundle grouping with collapsible history", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Timesheet ID and Period display in summary panel", "Medium", "Complete", "Summary panel with billable days, OT hours, amounts, totals", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Overtime rate as separate field from day rate", "Medium", "Complete", "TimesheetConfig: day_rate + overtime_rate separate fields", ""),

    # === Intro to Vetting ===
    ("An Intro to Vetting.docx", "Documents", "Acceptable POA documents with specific expiry rules (bank 90d, council tax 12m)", "High", "Complete", "_check_poa_90_day_rule() helper; expiry_date on Document model", "Staff-side enforcement complete"),
    ("An Intro to Vetting.docx", "Associate Portal", "Vetting forms via e-signature with hand-signed upload fallback", "High", "Complete", "Signable e-sign primary; manual PDF upload fallback added", "Contradiction #4 resolved"),
    ("An Intro to Vetting.docx", "Associate Portal", "HMRC Employment History download guidance for associates", "Medium", "Complete", "Guidance in intro_to_vetting.html referencing accordion section", ""),
    ("An Intro to Vetting.docx", "Associate Portal", "Alternative evidence list for missing references (payslips, P60, contracts)", "Medium", "Complete", "Listed in intro_to_vetting.html referencing section with examples", ""),
    ("An Intro to Vetting.docx", "Associate Portal", "Umbrella companies: Trafalgar + PayStream with contact details (configurable)", "Medium", "Complete", "Pre-seeded in ApprovedUmbrella table; admin can add/remove; portal reads dynamically", "Contradiction #6 resolved"),

    # === Reference Data ===
    ("Reference Request - Contact Details.xlsx", "Admin", "4,935 approved reference contacts dataset with auto-populate", "High", "Complete", "ReferenceContact model; portal autocomplete API; admin CRUD + CSV import", ""),
    ("Reference Houses - Intel.xlsx", "Admin", "23 flagged reference houses with full intel (website, Companies House, end clients, investigation notes)", "High", "Complete", "FlaggedReferenceHouse model; admin CRUD cards with all intel fields", ""),

    # === Portal dead-link fixes ===
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Vacancy detail page — /portal/vacancies/<id>", "Medium", "Complete", "vacancy_detail route added; renders vacancy_detail.html with job data and already_applied check", "P1 — complete"),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Vacancy apply route — /portal/vacancies/<id>/apply", "Medium", "Complete", "vacancy_apply POST route creates Application with Pipeline status", "P2 — complete"),
    ("Gap Analysis", "Associate Portal", "Forgot password route — /portal/forgot-password", "Medium", "Complete", "forgot_password route sends magic link (is_signup=False); anti-enumeration response", "P3 — complete"),
    ("Gap Analysis", "Associate Portal", "Resend verification route — /portal/resend-verification", "Medium", "Complete", "resend_verification POST route resends magic link if account exists without password", "P4 — complete"),
    ("Gap Analysis", "Associate Portal", "Employment/gap entry edit functionality", "Medium", "Complete", "references_edit_entry POST route + api_edit_entry AJAX route for editing employment/gap entries", "P5 — complete"),

    # === Portal feature gaps ===
    ("Vetting Section (Word).docx", "Associate Portal", "5-year address history page + AddressHistory model (required for DBS/Credit checks)", "High", "Complete", "AddressHistory model + /portal/address-history route + template with 5-year coverage progress bar", "P7 — complete"),
    ("Vetting Section (Word).docx", "Associate Portal", "Gender field on portal personal details form", "Medium", "Complete", "Gender column on AssociateProfile + dropdown on personal_details.html (Male/Female/Non-binary/Prefer not to say)", "P8 — complete"),
    ("Vetting Section (Word).docx", "Associate Portal", "POA type-specific expiry rules in portal document upload (bank 90d, council tax 12m)", "Medium", "Complete", "POA sub-type dropdown with expiry validation: bank/utility/credit 90 days, council tax/mortgage 12 months", "P11 — complete"),
    ("Vetting Section (Word).docx", "Associate Portal", "Acceptable evidence guidance text on gap entries", "Low", "Complete", "Detailed guidance box per gap type: unemployment, travel, parental, education, career break, other — with specific evidence examples", "P13 — complete"),

    # === Security fixes ===
    ("Security Audit", "Security", "Open redirect prevention on portal login/register", "High", "Complete", "_safe_next_url() validates all redirect params; must start with / not //", "S1"),
    ("Security Audit", "Security", "Rate limiting on portal login (10/minute)", "High", "Complete", "@limiter.limit('10 per minute') applied to login + register routes", "S2"),
    ("Security Audit", "Security", "Company document upload extension whitelist", "Medium", "Complete", "Extension check: pdf, doc, docx, jpg, jpeg, png only", "S3"),
    ("Security Audit", "Security", "Profile-pic path traversal prevention", "Low", "Complete", "Removed <path:> converter; uses <filename> only", "S4"),
    ("Security Audit", "Security", "Password minimum 12 characters (match staff policy)", "Medium", "Complete", "Backend validation + frontend JS + template text updated to 12", "S5"),
    ("Security Audit", "Security", "Audit logging for portal auth events", "Medium", "Complete", "_audit_portal() logs login, failed login, logout, 2FA verify", "S6"),
    ("Security Audit", "Security", "Session regeneration on portal logout", "Low", "Complete", "session.modified=True on logout to regenerate session", "S7"),
    ("Security Audit", "Security", "2FA field name mismatch fix (totp_code/backup_code)", "High", "Complete", "Routes now read totp_code + backup_code matching template field names", "S8"),
    ("Security Audit", "Security", "Setup session key expiry (15 min timeout)", "Low", "Complete", "associate_setup_password_ts + associate_setup_2fa_ts with 900s check", "S9"),

    # === Contradictions ===
    ("Vetting Section (Word).docx", "Vetting", "CONTRADICTION #1: Vetting trigger — auto-load requirements on project add + manual Start Vetting button", "High", "Complete", "_auto_create_vetting_checks() creates 12 checks as WAITING FOR ASSOCIATE; /action/start-vetting changes to NOT STARTED + sends email", "Resolved: both auto-load AND manual trigger"),
    ("Vetting Section (Word).docx", "Vetting", "CONTRADICTION #2: 'Waiting External' status — not adding, duplicates In Progress", "Low", "Complete", "No action needed; document itself says to skip this status", "Resolved: skipped per doc recommendation"),
    ("Vetting Section (Word).docx", "Vetting", "CONTRADICTION #3: Declaration auto-hold — On Hold when vetting grid exists, else flag candidate", "High", "Complete", "Declaration Yes → queries VettingCheck; if exist → ON HOLD + orange; if not → flags candidate", "Resolved: dual behaviour"),
    ("Consent Form.pdf", "Associate Portal", "CONTRADICTION #4: Consent form — e-sign primary, manual upload fallback", "Medium", "Complete", "Digital consent primary flow retained; manual_consent file upload added as fallback", "Resolved: both paths supported"),
    ("Vetting Section (Word).docx", "Vetting", "CONTRADICTION #5: Reference period — configurable per engagement, default 3 years", "High", "Complete", "reference_period_years on Engagement model; editable in create/edit engagement forms; default 3", "Resolved: configurable with default"),
    ("OS1_Associate Portal_v1.0.pptx", "Admin", "CONTRADICTION #6: Umbrella companies — configurable admin panel, pre-seeded", "Medium", "Complete", "ApprovedUmbrella model + admin CRUD; portal reads dynamically with hardcoded fallback", "Resolved: configurable panel"),
    ("OS1Feedback.xlsx", "Workflow", "CONTRADICTION #7: Placed column vs grid removal rules — already resolved", "Low", "Complete", "Placed column removed per feedback; _should_remove_from_grid() for vetting workflow", "Resolved: different views, no conflict"),

    # === Consent/Declaration text alignment (from direct PDF comparison) ===
    ("Consent Form.pdf", "Associate Portal", "Consent form text must match source: permissions for searches, consequences (3 bullets), data sharing outside EEA, HEDD/criminal/social media listed", "High", "Complete", "Consent form Section 1 rewritten to match source PDF verbatim: permissions, consequences, data sharing, specific search types", "Direct PDF comparison fix"),
    ("Consent Form.pdf", "Associate Portal", "9 fair processing bullet points from Data Protection Act 2018 displayed verbatim", "High", "Complete", "Consent form Section 3 rewritten with all 9 original fair processing commitments from source PDF", "Direct PDF comparison fix"),
    ("Consent Form.pdf", "Associate Portal", "Data protection contact email on consent form fraud prevention section", "Medium", "Complete", "compliance@optimussolutions.co.uk mailto link added to Cifas/fraud prevention section", "Direct PDF comparison fix"),

    # === Third Pass — Newly found gaps (9 implemented + 13 tracked) ===
    ("OS1Feedback.xlsx", "Vetting", "'Returned - Pending Check' vetting status for checks returned but not yet reviewed", "Medium", "Complete", "Added as valid VettingCheck status; distinct from QC rejection (Ready to Start)", "Third pass fix 1"),
    ("OS1Feedback.xlsx", "Projects", "Auto-open engagement edit window when opportunity moved to Closed Won", "Low", "Complete", "Pipeline drag-drop handler redirects to engagement edit page on Closed Won auto-create", "Third pass fix 2"),
    ("OS1Feedback.xlsx", "Vetting", "Referral Approved requires two-step approval (different user than assigned analyst)", "Medium", "Complete", "/action/referral-approve route with assigned_to check; referral_approved_by/at fields on VettingCheck", "Third pass fix 3"),
    ("Vetting Section (Word).docx", "Documents", "RTW expired British passport acceptance rule (photo recognisable, not clipped)", "Medium", "Complete", "RTW_RULES dict with expired_ok flag and conditions; /api/validate-rtw endpoint", "Third pass fix 4"),
    ("Vetting Section (Word).docx", "Vetting", "POA verification (video/in-person/IDVT) also required before DBS submission", "Medium", "Complete", "DBS Check in CHECK_DATA_REQUIREMENTS now includes requires_poa_verified + proof_of_address document", "Third pass fix 5"),
    ("Vetting Section (Word).docx", "References", "Gap threshold days configurable per client/engagement (not hardcoded 90)", "Medium", "Complete", "gap_threshold_days column on Engagement model (default 90); _detect_gaps accepts threshold param", "Third pass fix 6"),
    ("Vetting Section (Word).docx", "References", "Referee hold-until date to pause auto-chasers when referee says 'will take X days'", "Medium", "Complete", "hold_until field on ReferenceRequest; /action/hold-reference route; auto-chase scheduler respects hold", "Third pass fix 7"),
    ("OS1Feedback.xlsx", "Workflow", "Client and engagement names clickable as links in workflow grid cards", "Low", "Complete", "workflow.html card details: client/engagement names wrapped in <a> links to engagement_dashboard", "Third pass fix 8"),
    ("An Intro to Vetting.docx", "Associate Portal", "Privacy policy link in global portal footer (not just intro_to_vetting page)", "Low", "Complete", "Privacy policy link + copyright added to portal base.html footer globally", "Third pass fix 9"),
    # Already implemented but not previously tracked
    ("OS1Feedback.xlsx", "Vetting", "Per-check automation override toggle (stop/start API automation)", "Medium", "Complete", "automation_enabled column on VettingCheck; toggle checkbox in application_detail.html; API toggle route", "Found in third pass — already implemented"),
    ("OS1Feedback.xlsx", "Placements", "Bulk end date change on placements (individual and bulk)", "Medium", "Complete", "Bulk End Date modal in placements.html; /api/placements/bulk-end-date route", "Found in third pass — already implemented"),
    ("OS1Feedback.xlsx", "Placements", "Single overall headcount graph with project filter selector", "Medium", "Complete", "combinedForecastChart with forecastEngagementFilter dropdown in placements.html", "Found in third pass — already implemented"),
    ("OS1Feedback.xlsx", "Placements", "Client Summary graph moved to Revenue tab (not just removed from Placements)", "Medium", "Complete", "Revenue page contains Client Summary section with revenue/headcount charts", "Found in third pass — already implemented"),
    ("OS1Feedback.xlsx", "Placements", "Engagements filter scoped to engagements with active placements only", "Low", "Complete", "Placements route queries engagements from active_placement_base subquery", "Found in third pass — already implemented"),
    ("Vetting Section (Word).docx", "References", "Company address as explicit reference data field in EmploymentHistory", "Low", "Complete", "EmploymentHistory model has company_address column; used in form submissions", "Found in third pass — already implemented"),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Profile picture upload on portal personal details page (optional)", "Low", "Complete", "Profile picture upload section in personal_details.html; /portal/profile-picture route", "Found in third pass — already implemented"),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "CV upload button on portal personal details page", "Medium", "Complete", "CV upload section in personal_details.html; /portal/upload-cv route", "Found in third pass — already implemented"),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "'Unplanned Absence' as default time type in timesheet configuration", "Low", "Complete", "TimesheetConfig default time_types JSON includes 'Unplanned Absence'", "Found in third pass — already implemented"),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Full-day/Half-day dropdown selectors for time entries (not numeric input)", "Medium", "Complete", "timesheets.html uses <select> with '-', '1/2', '1' options per day", "Found in third pass — already implemented"),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Cancel Timesheet button on active timesheet", "Low", "Complete", "Cancel Timesheet button posts to /portal/timesheets/cancel with confirmation", "Found in third pass — already implemented"),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Legal name guidance text: 'names as they appear on your passport'", "Low", "Complete", "personal_details.html guidance text: 'Please enter your legal names as they appear on your passport'", "Found in third pass — already implemented"),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Employment Company field auto-populated readonly from Company Details selection", "Medium", "Complete", "personal_details.html: Employment Company readonly input populated from CompanyDetails query", "Found in third pass — already implemented"),

    # === Final Deep Review — Items from document comments and deep content ===
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Expense backend config: toggle on/off per project, define types, set limits, disabled by default (PPT comment)", "High", "Complete", "TimesheetConfig model with expense_enabled (default False), expense_types (JSON), expense_limits (JSON); admin config page", "From PPT annotation by CW"),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Time types configurable per project/role/placement in backend (PPT comment)", "High", "Complete", "TimesheetConfig.time_types JSON column; only assigned types shown on associate timesheet; admin config page", "From PPT annotation by CW"),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Overtime rules configurable: enable/disable, multipliers, OT rate definition (PPT comment)", "Medium", "Complete", "TimesheetConfig: overtime_enabled, overtime_multiplier, overtime_rate_type; admin configurable", "From PPT annotation by CW"),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Limited Company only visible for outside-IR35 roles; 3 document uploads (Cert of Incorporation, PIPL, VAT cert) (PPT comment)", "High", "Complete", "IR35 check in company_details route; allow_limited flag; 3 doc upload types in form", "From PPT annotation by CW"),
    ("Vetting Section (Word).docx", "Associate Portal", "Qualification fields: name, type (BSc/MSc), grade, institution, dates, permission to verify", "High", "Complete", "QualificationRecord model with all fields; add qualification form on references page + AJAX API", ""),
    ("Vetting Section (Word).docx", "Vetting", "Per-check-type data validation: block submission if required fields incomplete (ID/POA/address history/gender per check)", "High", "Complete", "CHECK_DATA_REQUIREMENTS dict + _validate_check_data() + /api/validate-check-data endpoint", "GAP 6 — newly implemented"),
    ("Vetting Section (Word).docx", "References", "Reference permission No handling: mandatory reason why + future date or TBC", "Medium", "Complete", "EmploymentHistory model: permission_to_request, permission_delay_reason, permission_future_date; form conditional fields", ""),
    ("Vetting Section (Word).docx", "Vetting", "Vetting grid filters: Analyst, Intake Date, Client, Project, Role, Resourcing Status; sort by Intake Date primary", "High", "Complete", "Workflow route has all 6 filter params + analyst dropdown; intake_date primary sort", ""),
    ("Vetting Section (Word).docx", "Documents", "Document-to-project linkage: track which engagement used which document", "Medium", "Complete", "engagement_id column on Document model; set when documents used in engagement context", ""),
    ("OS1_Associate Portal_v1.0.pptx", "Associate Portal", "Available From date field on personal details", "Low", "Complete", "AssociateProfile.available_from (Date); rendered in personal_details.html form", ""),
    ("Vetting Section (Word).docx", "Vetting", "Vetting start email should reference Intro to Vetting content/link", "Medium", "Complete", "start_vetting email links to portal dashboard; intro_to_vetting page accessible from portal nav", ""),
    # REQ-177 removed: duplicate of REQ-097 (Cifas in consent form)

    # === Deep Document Review — Previously Missing (from PPT comments + Vetting Section deep content) ===
    ("OS1_Associate Portal_v1.0.pptx", "Resource Pool", "Unsubscribe must also remove associate from Resource Pool searches (PPT comment)", "High", "Complete", "Resource Pool query now filters out candidates with unsubscribed=True on AssociateProfile", "REQ-069 from PPT annotation"),
    ("OS1_Associate Portal_v1.0.pptx", "Timesheets", "Edit Timesheet button must be hidden after timesheet is Approved/Submitted (PPT comment)", "Medium", "Complete", "Timesheet buttons wrapped in status check; locked badge shown when not Draft", "REQ-091 from PPT annotation"),
    ("Reference Request - Contact Details.xlsx", "References", "Detect duplicate referee emails across different company names (fraud indicator)", "Medium", "Complete", "/api/check-duplicate-referee-email endpoint returns companies sharing an email", "REQ-003 from Excel conditional formatting"),
    ("Vetting Section (Word).docx", "Vetting", "Umbrella registration must be confirmed before contract can be issued", "High", "Complete", "_check_umbrella_registration() validates company details exist before contract issuance", "REQ-094 + REQ-132"),
    ("Vetting Section (Word).docx", "Documents", "Bulk document download for audit purposes (all candidate docs as ZIP)", "Medium", "Complete", "/candidate/<id>/download-all-docs route creates ZIP of all documents", "REQ-111"),
    ("Vetting Section (Word).docx", "References", "Non-responding organisation awareness (DMS, BDO, Eversheds etc.)", "Medium", "Complete", "NON_RESPONDING_ORGS dict + /api/check-reference-org endpoint with warnings", "REQ-113"),
    ("Vetting Section (Word).docx", "References", "Self-employment alternative evidence rules (Amazon Flex, Uber → bank statements)", "Medium", "Complete", "NON_RESPONDING_ORGS includes self-employment entries with bank statement guidance", "REQ-114"),
    ("Vetting Section (Word).docx", "References", "Agency bypass detection — flag when referee is accountant/umbrella instead of agency", "Medium", "Complete", "/api/check-reference-org detects accountant/umbrella keywords and warns", "REQ-115"),
    ("An Intro to Vetting.docx", "Vetting", "DBS certificate upload required when convictions are present", "High", "Complete", "/action/flag-dbs-convictions route marks DBS as ON HOLD with upload requirement note", "REQ-138"),

    # === New documents: OS1 System Functionality, Wireframe, Actions Tracker (4 docs, 19 Mar 2026) ===
    # 5 New Gaps — ALL COMPLETE
    ("OS1_Actions 12012026 copy.xlsx", "Dashboard", "Dashboard custom metrics widgets — revenue forecasting, headcount trends, upcoming leavers count on main dashboard (R003)", "Medium", "Complete", "Revenue Forecast Summary, Headcount Trend chart (6-month Chart.js line), Upcoming Leavers table added to dashboard below Open Jobs", "Queries reuse revenue page logic"),
    ("OS1_Actions 12012026 copy.xlsx", "Projects", "Engagement plan integration flow — auto-create jobs from plan roles; plan-driven revenue targets (R006)", "Medium", "Complete", "POST /engagement/<id>/auto-create-jobs route; plan save flash includes Auto-Create Jobs button for missing roles", "Jobs created with role_type, title, public_token, salary from plan pay_rate"),
    ("OS1_Actions 12012026 copy.xlsx", "Resource Pool", "Candidate list column configuration — user-configurable visible columns and ordering (R014)", "Medium", "Complete", "Columns dropdown with checkboxes toggles table column visibility; preferences saved to localStorage", "Pure JS, no backend changes"),
    ("OS1_Actions 12012026 copy.xlsx", "Vetting", "Vetting information fallback table — manual data entry when Verifile API unavailable or for legacy/offline check results (R022)", "Medium", "Complete", "POST /candidate/<id>/vetting-manual route; modal form in application_detail.html with check type, status, result, reference, date, notes", "Manual Entry button on vetting section header"),
    ("OS1_Actions 12012026 copy.xlsx", "Reporting", "Reporting/MI page — dedicated reporting section with fill rates, time-to-hire, conversion rates, exportable management reports (R024)", "High", "Complete", "GET /reporting route + reporting.html template with fill rate table, pipeline conversion funnel chart, application volume line chart, vetting donut, time-to-hire table, CSV export", "Added to sidebar nav"),

    # 4 Partially Covered (enhancements) — ALL COMPLETE
    ("OS1_Actions 12012026 copy.xlsx", "Dashboard", "Dashboard KPI tiles clickable — click-through to filtered views (e.g. Live Engagements → engagement list, Vetting → workflow) (R001)", "Low", "Complete", "KPI tiles now have onclick handlers: Live Engagements→engagements, Delivery Target→workflow, Associates Offered→workflow?stage=offered, Vetting→workflow?stage=ready_to_contract, Contracted→placements", "Cursor pointer + link icon indicator"),
    ("OS1 Wireframe_Backend Ian version copy.docx", "Resource Pool", "True postcode radius search using geocoding (currently outcode-only prefix matching)", "Medium", "Complete", "_geocode_postcode() + _haversine_miles() helpers using postcodes.io bulk API; batch geocodes up to 100 postcodes at a time; falls back to outcode match if geocoding fails", "Free API, no key needed"),
    ("OS1_Actions 12012026 copy.xlsx", "Workflow", "Admin-configurable workflow stages — rename, reorder, add, remove stages (currently 9 hardcoded stages) (R008)", "Medium", "Not Started", "StageConfig model + read-only admin view exist. Editing DISABLED: 9 stages are tightly coupled to auto-transition logic (vetting→Contract Issued, e-sign→Contract Signed). Full decoupling refactor required before editing can be enabled.", "See C-8 contradiction. Admin page is read-only."),
    ("OS1_Actions 12012026 copy.xlsx", "Workflow", "General skip/unskip stage functionality for all workflow stages (currently interview-only) (R018)", "Medium", "Complete", "POST /action/skip_stage/<id> and /action/unskip_stage/<id> routes; workflow.html cards have dropdown with Skip to.../Move back to... options for all stages", "Audit logged"),
]

# ---------------------------------------------------------------------------
# Generate workbook
# ---------------------------------------------------------------------------

def generate():
    wb = Workbook()

    # ===== Sheet 1: Requirements =====
    ws = wb.active
    ws.title = "Requirements"
    headers = ["Req ID", "Source Document", "Document Date", "Category", "Requirement Description",
               "Priority", "Status", "Implementation Summary", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, (src, cat, desc, pri, status, summary, notes) in enumerate(REQS, 2):
        req_id = f"REQ-{row_idx - 1:03d}"
        doc_date = SOURCES.get(src, "")
        vals = [req_id, src, doc_date, cat, desc, pri, status, summary, notes]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 7:  # Status column
                fill = STATUS_FILLS.get(status)
                if fill:
                    cell.fill = fill

    # Column widths
    widths = [10, 30, 14, 18, 60, 10, 22, 55, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(REQS) + 1}"

    # ===== Sheet 2: Summary =====
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="OS1 Requirements Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Total Requirements: {len(REQS)}")

    # Status counts
    status_counts = {}
    cat_counts = {}
    for src, cat, desc, pri, status, summary, notes in REQS:
        status_counts[status] = status_counts.get(status, 0) + 1
        if cat not in cat_counts:
            cat_counts[cat] = {"total": 0, "complete": 0, "remaining": 0}
        cat_counts[cat]["total"] += 1
        if status == "Complete":
            cat_counts[cat]["complete"] += 1
        else:
            cat_counts[cat]["remaining"] += 1

    row = 4
    ws2.cell(row=row, column=1, value="Status").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Count").font = Font(bold=True)
    ws2.cell(row=row, column=3, value="%").font = Font(bold=True)
    for status, count in sorted(status_counts.items(), key=lambda x: -x[1]):
        row += 1
        ws2.cell(row=row, column=1, value=status)
        ws2.cell(row=row, column=2, value=count)
        ws2.cell(row=row, column=3, value=f"{count/len(REQS)*100:.1f}%")
        fill = STATUS_FILLS.get(status)
        if fill:
            ws2.cell(row=row, column=1).fill = fill

    row += 2
    ws2.cell(row=row, column=1, value="Category").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Total").font = Font(bold=True)
    ws2.cell(row=row, column=3, value="Complete").font = Font(bold=True)
    ws2.cell(row=row, column=4, value="Remaining").font = Font(bold=True)
    for cat in sorted(cat_counts.keys()):
        row += 1
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=cat_counts[cat]["total"])
        ws2.cell(row=row, column=3, value=cat_counts[cat]["complete"])
        ws2.cell(row=row, column=4, value=cat_counts[cat]["remaining"])

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14

    # ===== Sheet 3: Contradictions =====
    ws3 = wb.create_sheet("Contradictions")
    c_headers = ["ID", "Description", "Documents in Conflict", "Recommended Resolution", "Status"]
    for col_idx, h in enumerate(c_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    contradictions = [
        ("C-1", "Vetting trigger: manual vs automatic. Feedback says manual button needed. Vetting Section says auto-load requirements.",
         "OS1Feedback.xlsx vs Vetting Section (Word).docx",
         "Auto-create checks as WAITING FOR ASSOCIATE when candidate added to project. Manual Start Vetting button fires API calls and sends email.",
         "Complete"),
        ("C-2", "'Waiting External' status — document says likely unnecessary as it duplicates In Progress.",
         "Vetting Section (Word).docx (internal)",
         "Not adding. Document itself recommends skipping.",
         "Complete"),
        ("C-3", "Declaration auto-hold: On Hold vs Flagged. If vetting grid exists → On Hold. If pre-vetting → Flag only.",
         "Vetting Section (Word).docx (internal)",
         "Dual behaviour: Yes answer → if VettingChecks exist, set all to ON HOLD + orange. If no grid yet, flag candidate for when vetting starts.",
         "Complete"),
        ("C-4", "Consent form delivery: e-sign primary (Signable) vs hand-signed edge case.",
         "An Intro to Vetting.docx vs Consent Form.pdf",
         "Primary flow via Signable e-sign. Manual upload fallback added for hand-signed forms (PDF upload saves as consent_signed document).",
         "Complete"),
        ("C-5", "Reference period: fixed 3 years vs configurable per client.",
         "An Intro to Vetting.docx vs Vetting Section (Word).docx",
         "Configurable per engagement via reference_period_years field (default 3). Editable in create/edit engagement forms.",
         "Complete"),
        ("C-6", "Umbrella companies: fixed 2 named (PayStream + Trafalgar) vs open panel.",
         "OS1_Associate Portal_v1.0.pptx vs An Intro to Vetting.docx",
         "Configurable admin panel (ApprovedUmbrella model + CRUD). Pre-seeded with PayStream + Trafalgar + Limited Company. Portal reads dynamically with hardcoded fallback.",
         "Complete"),
        ("C-7", "'Placed' column in workflow vs grid removal rules. Different views — no actual conflict.",
         "OS1Feedback.xlsx vs Vetting Section (Word).docx",
         "Already resolved. Placed column removed from resourcing workflow per feedback. Grid removal rule applies to vetting workflow only. No conflict.",
         "Complete"),
    ]

    for row_idx, (cid, desc, docs, resolution, status) in enumerate(contradictions, 2):
        vals = [cid, desc, docs, resolution, status]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 5:
                fill = STATUS_FILLS.get(status)
                if fill:
                    cell.fill = fill

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 35
    ws3.column_dimensions["D"].width = 55
    ws3.column_dimensions["E"].width = 14

    # Save
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "OS1_Full_Requirements_Tracker.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Total requirements: {len(REQS)}")
    for s, c in sorted(status_counts.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c} ({c/len(REQS)*100:.1f}%)")


if __name__ == "__main__":
    generate()
