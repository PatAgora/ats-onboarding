"""
Import reference contact data and flagged reference houses into the database.
Run once after the associate portal models are created.

Usage:
    python import_reference_data.py
"""
import os
import sys
import datetime

# Add parent dir so we can import from app
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def import_reference_contacts():
    """Import 4,935 known reference contacts from Excel."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed, skipping reference contact import")
        return 0

    xlsx_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..",
        "Reference Request - Contact Details.xlsx",
    )
    if not os.path.exists(xlsx_path):
        print(f"Reference contacts file not found: {xlsx_path}")
        return 0

    from app import engine, Base
    from associate_portal import ReferenceContact
    from sqlalchemy.orm import Session
    from sqlalchemy import select

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    count = 0
    with Session(engine) as s:
        # Check if already imported
        existing = s.scalar(select(ReferenceContact.id).limit(1))
        if existing:
            print("Reference contacts already imported, skipping.")
            return 0

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
        for row in rows:
            if not row or not row[0]:
                continue
            company_name = str(row[0]).strip()
            referee_email = str(row[1]).strip() if row[1] else ""
            last_amended = row[2] if isinstance(row[2], datetime.datetime) else None

            if company_name and referee_email:
                contact = ReferenceContact(
                    company_name=company_name,
                    referee_email=referee_email,
                    last_amended=last_amended,
                )
                s.add(contact)
                count += 1

        s.commit()
    wb.close()
    print(f"Imported {count} reference contacts.")
    return count


def import_flagged_reference_houses():
    """Import 23 known fake/suspicious reference houses from Excel."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed, skipping reference house import")
        return 0

    xlsx_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..",
        "Reference Houses - Intel.xlsx",
    )
    if not os.path.exists(xlsx_path):
        print(f"Reference houses file not found: {xlsx_path}")
        return 0

    from app import engine, Base
    from associate_portal import FlaggedReferenceHouse
    from sqlalchemy.orm import Session
    from sqlalchemy import select

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    count = 0
    with Session(engine) as s:
        # Check if already imported
        existing = s.scalar(select(FlaggedReferenceHouse.id).limit(1))
        if existing:
            print("Flagged reference houses already imported, skipping.")
            return 0

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
        for row in rows:
            if not row or not row[1]:  # Column B has the name
                continue
            name = str(row[1]).strip()
            candidate_count = int(row[2]) if row[2] else 0
            end_clients = str(row[3]).strip() if row[3] else ""
            website = str(row[4]).strip() if row[4] else ""
            companies_house_url = str(row[5]).strip() if row[5] else ""
            notes = str(row[6]).strip() if row[6] else ""

            if name:
                house = FlaggedReferenceHouse(
                    name=name,
                    candidate_count=candidate_count,
                    end_clients=end_clients,
                    website=website,
                    companies_house_url=companies_house_url,
                    notes=notes,
                )
                s.add(house)
                count += 1

        s.commit()
    wb.close()
    print(f"Imported {count} flagged reference houses.")
    return count


if __name__ == "__main__":
    contacts = import_reference_contacts()
    houses = import_flagged_reference_houses()
    print(f"\nDone. {contacts} contacts, {houses} flagged houses imported.")
